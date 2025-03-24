import asyncio
import pandas as pd
from datetime import datetime
from pathlib import Path
import shutil
import logging
import aiohttp  # Add this import
from parsers.kwork_parser import KworkParser
from parsers.freelancehunt_parser import FreelancehuntParser
from parsers.flru_parser import FLRuParser
from config import EXCEL_FILE, PARSING_INTERVAL
from openpyxl import load_workbook  # Changed this line
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo  # Add this import

# Add logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('parser.log'),
        logging.StreamHandler()
    ]
)

# Add backup functionality
def backup_data():
    backup_dir = Path('data/backups')
    backup_dir.mkdir(parents=True, exist_ok=True)
    
    if Path(EXCEL_FILE).exists():
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = backup_dir / f'freelance_jobs_backup_{timestamp}.xlsx'
        shutil.copy2(EXCEL_FILE, backup_path)
        
        # Keep only last 5 backups
        backups = sorted(backup_dir.glob('*.xlsx'))
        if len(backups) > 5:
            for old_backup in backups[:-5]:
                old_backup.unlink()

def format_excel(filename, new_entries_count=0):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Format headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Add borders
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # Title
    ws.column_dimensions['B'].width = 60  # Description
    ws.column_dimensions['C'].width = 15  # Price
    ws.column_dimensions['D'].width = 20  # Client
    ws.column_dimensions['E'].width = 30  # URL
    ws.column_dimensions['F'].width = 20  # Date
    ws.column_dimensions['G'].width = 15  # Platform
    
    # Format data cells and highlight new entries
    new_entry_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_rows = ws.max_row
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Highlight new entries
            if row[0].row > (total_rows - new_entries_count):
                cell.fill = new_entry_fill
            
            # Format price column (column C)
            if cell.column_letter == 'C' and cell.value:
                cell.number_format = '"$"#,##0.00_);("$"#,##0.00)'
            
            # Format date column (column F)
            if cell.column_letter == 'F' and cell.value:
                cell.number_format = 'yyyy-mm-dd hh:mm'
    
    # Add table for easy filtering and sorting
    tab = Table(displayName="JobsTable", ref=f"A1:G{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    wb.save(filename)

async def main():
    # Create data directory if it doesn't exist
    Path(EXCEL_FILE).parent.mkdir(parents=True, exist_ok=True)
    
    parsers = [
        KworkParser(),
        FreelancehuntParser(),
        FLRuParser(),
        # More parsers will be added here
    ]
    
    while True:
        all_jobs = []
        
        for parser in parsers:
            try:
                jobs = await parser.parse()
                all_jobs.extend(jobs)
                logging.info(f"Found {len(jobs)} new jobs from {parser.platform_name}")
            except aiohttp.ClientError as e:
                logging.error(f"Network error for {parser.platform_name}: {str(e)}")
            except Exception as e:
                logging.error(f"Error parsing {parser.platform_name}: {str(e)}")
        
        if all_jobs:
            # Convert to DataFrame
            df_new = pd.DataFrame([
                {
                    'Title': job.title,
                    'Description': job.description,
                    'Price': job.price,
                    'Client': job.client_name,
                    'URL': job.url,
                    'Date': job.published_date,
                    'Platform': job.platform
                }
                for job in all_jobs
            ])
            
            # Load existing data or create new file
            if Path(EXCEL_FILE).exists():
                df_existing = pd.read_excel(EXCEL_FILE)
                df = pd.concat([df_existing, df_new]).drop_duplicates(subset=['URL'])
            else:
                df = df_new
            
            # Create backup before saving new data
            backup_data()
            
            # Save to Excel and format
            df.to_excel(EXCEL_FILE, index=False)
            format_excel(EXCEL_FILE, len(df_new))
            
            # Log summary
            logging.info(f"Saved {len(df_new)} new jobs. Total jobs: {len(df)}")
        
        await asyncio.sleep(PARSING_INTERVAL * 60)

if __name__ == "__main__":
    print("Starting freelance job parser...")
    asyncio.run(main())